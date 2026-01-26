import os
import io
import json
import locale
import openpyxl 
from copy import copy 
from datetime import date, datetime
from django.utils import timezone 
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Max, Q, Sum
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors

from gestion.models import Trabajador, DepositoDiario, DepositoAporte, DepositoDesglose, BODEGA_FILTRO_CHOICES

try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass

# =========================================================================
# VISTAS GENERALES DEL BANCO
# =========================================================================

@login_required
def total_dia(request):
    return render(request, 'gestion/caja_banco/banco_menu.html')

@login_required
def resumen_consolidado(request):
    fecha_str_seleccionada = request.GET.get('fecha_seleccionada')
    if fecha_str_seleccionada:
        try:
            fecha_seleccionada = date.fromisoformat(fecha_str_seleccionada)
        except ValueError:
            fecha_seleccionada = date.today()
    else:
        fecha_seleccionada = date.today()

    # Lógica de Totales
    totales_bodega = DepositoDiario.objects.filter(
        fecha=fecha_seleccionada
    ).values(
        'bodega_nombre'
    ).annotate(
        total_aportes=Sum('total_aportes'),
        total_desglose=Sum('total_desglose'),
        total_cheques=Sum('total_cheques'),
        total_diferencia=Sum('diferencia')
    ).order_by('bodega_nombre')

    datos_mp = {'aportes': 0, 'desglose_efectivo': 0, 'cheques': 0, 'desglose_total': 0, 'diferencia': 0}
    datos_dp = {'aportes': 0, 'desglose_efectivo': 0, 'cheques': 0, 'desglose_total': 0, 'diferencia': 0}

    for data in totales_bodega:
        total_general_desglose = (data['total_desglose'] or 0) + (data['total_cheques'] or 0)
        
        info = {
            'aportes': data['total_aportes'] or 0,
            'desglose_efectivo': data['total_desglose'] or 0,
            'cheques': data['total_cheques'] or 0,
            'desglose_total': total_general_desglose,
            'diferencia': data['total_diferencia'] or 0
        }

        if data['bodega_nombre'] == 'Manuel Peñafiel':
            datos_mp = info
        elif 'David' in data['bodega_nombre']:
            datos_dp = info

    total_consolidado_data = {
        'aportes': datos_mp['aportes'] + datos_dp['aportes'],
        'desglose_total': datos_mp['desglose_total'] + datos_dp['desglose_total'],
        'diferencia': datos_mp['diferencia'] + datos_dp['diferencia']
    }

    # Listas
    aportes_mp_list = DepositoAporte.objects.filter(
        deposito__fecha=fecha_seleccionada,
        deposito__bodega_nombre='Manuel Peñafiel'
    ).values('trabajador__nombre').annotate(total=Sum('monto')).order_by('trabajador__nombre')

    aportes_dp_list = DepositoAporte.objects.filter(
        deposito__fecha=fecha_seleccionada,
        deposito__bodega_nombre='David Perry'
    ).values('trabajador__nombre').annotate(total=Sum('monto')).order_by('trabajador__nombre')

    # Desglose Billetes
    desglose_query = DepositoDesglose.objects.filter(
        deposito__fecha=fecha_seleccionada
    ).values('valor_unitario').annotate(
        cantidad_total=Sum('cantidad'),
        monto_total=Sum('total_denominacion')
    ).order_by('-valor_unitario')

    denominaciones_std = [20000, 10000, 5000, 2000, 1000, 500, 100, 50, 10]
    desglose_final = []
    mapa_desglose = {d['valor_unitario']: d for d in desglose_query}

    for valor in denominaciones_std:
        data = mapa_desglose.get(valor, {'cantidad_total': 0, 'monto_total': 0})
        desglose_final.append({
            'texto': f"${valor:,}".replace(",", "."),
            'cantidad': data['cantidad_total'],
            'total': data['monto_total']
        })

    total_cheques_dia = datos_mp['cheques'] + datos_dp['cheques']

    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    dia_sem = dias_semana[fecha_seleccionada.weekday()]
    mes_nom = meses[fecha_seleccionada.month]
    fecha_bonita = f"{dia_sem}, {fecha_seleccionada.day} de {mes_nom} de {fecha_seleccionada.year}"

    context = {
        'mp': datos_mp,
        'dp': datos_dp,
        'aportes_mp_list': aportes_mp_list,
        'aportes_dp_list': aportes_dp_list,
        'total': total_consolidado_data,
        'desglose_final': desglose_final,
        'total_cheques_dia': total_cheques_dia,
        'fecha_seleccionada_str': fecha_seleccionada.strftime('%Y-%m-%d'),
        'fecha_bonita': fecha_bonita
    }
    
    return render(request, 'gestion/caja_banco/banco_total_dia.html', context)

@login_required
def deposito_bodega(request, bodega_nombre):
    fecha_str_seleccionada = request.GET.get('fecha_seleccionada')
    
    if fecha_str_seleccionada:
        try:
            fecha_seleccionada = date.fromisoformat(fecha_str_seleccionada)
        except ValueError:
            fecha_seleccionada = date.today()
    else:
        fecha_seleccionada = date.today()
    
    lotes_del_dia = DepositoDiario.objects.filter(
        fecha=fecha_seleccionada, 
        bodega_nombre=bodega_nombre
    ).order_by('numero_lote')
    
    context = {
        'bodega_nombre': bodega_nombre,
        'lotes_del_dia': lotes_del_dia,
        'fecha_seleccionada_str': fecha_seleccionada.strftime('%Y-%m-%d'),
    }
    return render(request, 'gestion/caja_banco/banco_lotes/banco_lotes.html', context)

@login_required
def reportes_mensuales(request):
    bodegas_opciones = ["TODAS (Consolidado)"] + [choice[0] for choice in BODEGA_FILTRO_CHOICES]
    meses_opciones = [
        ("1", "Enero"), ("2", "Febrero"), ("3", "Marzo"), ("4", "Abril"),
        ("5", "Mayo"), ("6", "Junio"), ("7", "Julio"), ("8", "Agosto"),
        ("9", "Septiembre"), ("10", "Octubre"), ("11", "Noviembre"), ("12", "Diciembre")
    ]
    años_opciones = list(range(2025, date.today().year + 2))
    
    today = date.today()
    bodega_sel = request.GET.get('bodega', 'TODAS (Consolidado)')
    mes_sel = request.GET.get('mes', str(today.month))
    año_sel = request.GET.get('año', str(today.year))

    queryset = DepositoDiario.objects.filter(
        fecha__year=int(año_sel),
        fecha__month=int(mes_sel)
    )
    
    if bodega_sel != 'TODAS (Consolidado)':
        queryset = queryset.filter(bodega_nombre=bodega_sel)
        
    reporte_data = queryset.order_by('fecha', 'bodega_nombre', 'numero_lote')
    
    totales = reporte_data.aggregate(
        total_aportes=Sum('total_aportes'),
        total_desglose=Sum('total_desglose'),
        total_cheques=Sum('total_cheques'),
        total_diferencia=Sum('diferencia')
    )
    
    total_desglose_general = (totales['total_desglose'] or 0) + (totales['total_cheques'] or 0)
    total_aportes_general = (totales['total_aportes'] or 0)
    diferencia_real_total = total_aportes_general - total_desglose_general

    context = {
        'reporte_data': reporte_data,
        'total_aportes': total_aportes_general,
        'total_desglose_general': total_desglose_general,
        'diferencia_real_total': diferencia_real_total,
        'bodegas_opciones': bodegas_opciones,
        'meses_opciones': meses_opciones,
        'años_opciones': años_opciones,
        'bodega_sel': bodega_sel,
        'mes_sel': mes_sel,
        'año_sel': int(año_sel),
    }
    return render(request, 'gestion/caja_banco/banco_mensual.html', context)

# =========================================================================
# GESTIÓN DE LOTES
# =========================================================================

@login_required
def crear_nuevo_lote(request):
    if request.method == 'POST':
        bodega = request.POST.get('bodega_nombre')
        fecha_str = request.POST.get('fecha_seleccionada')
        nombre_lote = request.POST.get('nombre_lote', '')
        
        fecha_obj = date.fromisoformat(fecha_str)

        resultado = DepositoDiario.objects.filter(
            fecha=fecha_obj, 
            bodega_nombre=bodega
        ).aggregate(max_lote=Max('numero_lote'))
        
        nuevo_numero = (resultado['max_lote'] or 0) + 1
        
        DepositoDiario.objects.create(
            fecha=fecha_obj,
            bodega_nombre=bodega,
            numero_lote=nuevo_numero,
            nombre_lote=nombre_lote
        )

        url_redireccion = f"{reverse('deposito_bodega', args=[bodega])}?fecha_seleccionada={fecha_str}"
        return redirect(url_redireccion)
    
    return redirect('total_dia')

@login_required
def editar_lote(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    puede_editar = not lote.cerrado

    denominaciones = [
        {"texto": "$20.000", "valor": 20000}, {"texto": "$10.000", "valor": 10000},
        {"texto": "$5.000", "valor": 5000}, {"texto": "$2.000", "valor": 2000},
        {"texto": "$1.000", "valor": 1000}, {"texto": "$500", "valor": 500},
        {"texto": "$100", "valor": 100}, {"texto": "$50", "valor": 50},
        {"texto": "$10", "valor": 10},
    ]

    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'reabrir_lote' and request.user.is_superuser:
            lote.cerrado = False
            lote.save() 
            return redirect('editar_lote', lote_id=lote.id)

        if not puede_editar:
            return redirect('editar_lote', lote_id=lote.id)

        total_desglose_efectivo = 0
        lote.desglose.all().delete() 
        
        for d in denominaciones:
            cantidad_str = request.POST.get(f"cant_{d['valor']}")
            cantidad = int(cantidad_str) if cantidad_str else 0
            
            if cantidad > 0:
                total_denominacion = cantidad * d['valor']
                total_desglose_efectivo += total_denominacion
                
                DepositoDesglose.objects.create(
                    deposito=lote,
                    denominacion=d['texto'],
                    valor_unitario=d['valor'],
                    cantidad=cantidad,
                    total_denominacion=total_denominacion
                )
        
        cheque_str = request.POST.get('cant_cheque', '0')
        cheque_monto = int(cheque_str) if cheque_str else 0
        
        lote.total_desglose = total_desglose_efectivo
        lote.total_cheques = cheque_monto
        
        total_desglose_general = total_desglose_efectivo + cheque_monto
        lote.diferencia = lote.total_aportes - total_desglose_general
        
        if accion == 'cerrar_lote':
            lote.cerrado = True
        
        lote.save() 
        return redirect('editar_lote', lote_id=lote.id)

    # Lógica de filtrado de trabajadores
    if 'Manuel' in lote.bodega_nombre: 
        filtro_bodega = '1221'  
    else:
        filtro_bodega = '1225'  
        
    trabajadores_disponibles = Trabajador.objects.filter(
        Q(bodega_asignada=filtro_bodega) | Q(bodega_asignada='Ambos'),
        activo=True,              
        filtro_trabajador=True    
    ).order_by('nombre')

    aportes = lote.aportes.all().order_by('trabajador__nombre')
    desglose_guardado = {d.valor_unitario: d.cantidad for d in lote.desglose.all()}
    
    desglose_para_plantilla = []
    for d in denominaciones:
        cantidad_guardada = desglose_guardado.get(d['valor'], 0)
        desglose_para_plantilla.append({
            "texto": d['texto'],
            "valor": d['valor'],
            "cantidad": cantidad_guardada,
            "total_formateado": f"${cantidad_guardada * d['valor']:,}".replace(",", ".")
        })

    context = {
        'lote': lote,
        'aportes': aportes,
        'trabajadores_disponibles': trabajadores_disponibles,
        'denominaciones': desglose_para_plantilla,
        'cheque_guardado': lote.total_cheques,
        'lote_esta_cerrado': lote.cerrado,
        'puede_editar': puede_editar
    }
    return render(request, 'gestion/caja_banco/banco_lotes/banco_lotes_editor.html', context)

@login_required
def agregar_aporte(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    
    if request.method == 'POST':
        trabajador_id = request.POST.get('trabajador')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion', '')
        
        if trabajador_id and monto:
            trabajador = get_object_or_404(Trabajador, id=trabajador_id)
            monto_int = int(monto)

            DepositoAporte.objects.create(
                deposito=lote,
                trabajador=trabajador,
                monto=monto_int,
                descripcion=descripcion
            )
            
            lote.total_aportes += monto_int
            total_desglose_general = lote.total_desglose + lote.total_cheques
            lote.diferencia = lote.total_aportes - total_desglose_general
            lote.save()
    
    return redirect('editar_lote', lote_id=lote.id)

@login_required
def quitar_aportes_seleccion(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    
    if lote.cerrado:
        return redirect('editar_lote', lote_id=lote.id)

    if request.method == 'POST':
        lista_ids_aportes = request.POST.getlist('aporte_ids')
        
        if lista_ids_aportes:
            aportes_a_borrar = DepositoAporte.objects.filter(id__in=lista_ids_aportes, deposito=lote)
            total_restando = aportes_a_borrar.aggregate(total=Sum('monto'))['total'] or 0
            
            lote.total_aportes -= total_restando
            total_desglose_general = lote.total_desglose + lote.total_cheques
            lote.diferencia = lote.total_aportes - total_desglose_general
            
            lote.save()
            aportes_a_borrar.delete()

        return redirect('editar_lote', lote_id=lote.id)
    
    return redirect('total_dia')

@login_required 
def desbloquear_lote(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)

    if not request.user.is_superuser:
        return redirect('home')
        
    lote.cerrado = False
    lote.save()
    return redirect('editar_lote', lote_id=lote.id)

@login_required
def eliminar_lote(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    
    if not request.user.is_superuser and lote.cerrado:
        return redirect('home')  

    if request.method == 'POST':
        bodega_nombre = lote.bodega_nombre
        fecha_str = lote.fecha.strftime('%Y-%m-%d')
        lote.delete()
        url_redireccion = f"{reverse('deposito_bodega', args=[bodega_nombre])}?fecha_seleccionada={fecha_str}"
        return redirect(url_redireccion)
    
    return redirect('total_dia')

@login_required
def renombrar_lote(request, lote_id):
    if request.method == 'POST':
        lote = get_object_or_404(DepositoDiario, id=lote_id)
        nuevo_nombre = request.POST.get('nuevo_nombre')
        
        if not request.user.is_superuser and lote.cerrado:
            return redirect('deposito_bodega', bodega_nombre=lote.bodega_nombre)

        lote.nombre_lote = nuevo_nombre
        lote.save()
        return redirect('editar_lote', lote_id=lote.id)
    
    return redirect('home')

@login_required
def editar_aporte(request, aporte_id):
    aporte = get_object_or_404(DepositoAporte, id=aporte_id)
    lote = aporte.deposito
    
    if lote.cerrado and not request.user.is_superuser:
        return redirect('editar_lote', lote_id=lote.id)

    if request.method == 'POST':
        trabajador_id = request.POST.get('trabajador')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion', '')

        if trabajador_id and monto:
            lote.total_aportes -= aporte.monto
            aporte.trabajador_id = trabajador_id
            aporte.monto = int(monto)
            aporte.descripcion = descripcion
            aporte.save()
            lote.total_aportes += aporte.monto
            total_desglose_general = lote.total_desglose + lote.total_cheques
            lote.diferencia = lote.total_aportes - total_desglose_general
            lote.save()

    return redirect('editar_lote', lote_id=lote.id)

@login_required
def api_auto_guardar_arqueo(request, lote_id):
    if request.method == 'POST':
        lote = get_object_or_404(DepositoDiario, id=lote_id)
        
        if lote.cerrado and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Lote cerrado'}, status=403)

        denominaciones = [20000, 10000, 5000, 2000, 1000, 500, 100, 50, 10]
        mapa_textos = {
            20000: "$20.000", 10000: "$10.000", 5000: "$5.000", 2000: "$2.000",
            1000: "$1.000", 500: "$500", 100: "$100", 50: "$50", 10: "$10"
        }

        total_desglose_efectivo = 0
        lote.desglose.all().delete()
        
        for valor in denominaciones:
            cantidad_str = request.POST.get(f"cant_{valor}")
            cantidad = int(cantidad_str) if cantidad_str else 0
            
            if cantidad > 0:
                total_denominacion = cantidad * valor
                total_desglose_efectivo += total_denominacion
                DepositoDesglose.objects.create(
                    deposito=lote,
                    denominacion=mapa_textos.get(valor, f"${valor}"),
                    valor_unitario=valor,
                    cantidad=cantidad,
                    total_denominacion=total_denominacion
                )
        
        cheque_str = request.POST.get('cant_cheque', '0')
        cheque_monto = int(cheque_str) if cheque_str else 0
        
        lote.total_desglose = total_desglose_efectivo
        lote.total_cheques = cheque_monto
        
        total_desglose_general = total_desglose_efectivo + cheque_monto
        lote.diferencia = lote.total_aportes - total_desglose_general
        lote.save()
        return JsonResponse({'status': 'ok', 'message': 'Guardado'})
    
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def exportar_lote_excel(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    plantilla_path = os.path.join(settings.BASE_DIR, 'plantillas', 'PlantillaBanco.xlsx')
    
    try:
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active
    except FileNotFoundError:
        return HttpResponse("Error: PlantillaBanco.xlsx no encontrada.", status=404)

    ws['B2'] = f"Lote N°{lote.numero_lote}"
    ws['D2'] = lote.bodega_nombre
    ws['H2'] = lote.fecha.strftime('%d-%m-%Y')
    
    if lote.updated_at:
        fecha_local = timezone.localtime(lote.updated_at)
        ws['K2'] = f"Cerrado a las: {fecha_local.strftime('%H:%M:%S')}"
    else:
        ws['K2'] = "No cerrado"

    fila_map = {20000: 7, 10000: 8, 5000: 9, 2000: 10, 1000: 11, 500: 12, 100: 13, 50: 14, 10: 15}
    for d in lote.desglose.all():
        fila = fila_map.get(d.valor_unitario)
        if fila:
            ws[f'K{fila}'] = d.cantidad
            ws[f'M{fila}'] = d.total_denominacion

    ws['L17'] = lote.total_desglose

    aportes = lote.aportes.all().order_by('id')
    base_row = 7 
    
    for i, aporte in enumerate(aportes):
        current_row = base_row + (i * 2)
        if i > 0:
            for col in range(2, 8): 
                source_top = ws.cell(row=7, column=col)
                target_top = ws.cell(row=current_row, column=col)
                if source_top.has_style:
                    target_top.font = copy(source_top.font)
                    target_top.border = copy(source_top.border)
                    target_top.fill = copy(source_top.fill)
                    target_top.alignment = copy(source_top.alignment)
                    target_top.number_format = copy(source_top.number_format)
                
                source_btm = ws.cell(row=8, column=col)
                target_btm = ws.cell(row=current_row + 1, column=col)
                if source_btm.has_style:
                    target_btm.font = copy(source_btm.font)
                    target_btm.border = copy(source_btm.border)
                    target_btm.fill = copy(source_btm.fill)
                    target_btm.alignment = copy(source_btm.alignment)
                    target_btm.number_format = copy(source_btm.number_format)

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)     
            ws.merge_cells(start_row=current_row+1, start_column=2, end_row=current_row+1, end_column=5) 
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row+1, end_column=7)   

        ws.cell(row=current_row, column=2).value = aporte.trabajador.nombre
        ws.cell(row=current_row+1, column=2).value = aporte.descripcion
        ws.cell(row=current_row, column=6).value = aporte.monto 

    num_aportes = len(aportes)
    fila_total = base_row + (num_aportes * 2) + 1
    
    source_label = ws['I17']
    target_label = ws.cell(row=fila_total, column=2) 
    if source_label.has_style:
        target_label.font = copy(source_label.font)
        target_label.border = copy(source_label.border)
        target_label.fill = copy(source_label.fill)
        target_label.alignment = copy(source_label.alignment)
        target_label.number_format = copy(source_label.number_format)
        
    target_label.value = "TOTAL APORTES"
    ws.merge_cells(start_row=fila_total, start_column=2, end_row=fila_total, end_column=5)

    source_val = ws['L17']
    target_val = ws.cell(row=fila_total, column=6) 
    if source_val.has_style:
        target_val.font = copy(source_val.font)
        target_val.border = copy(source_val.border)
        target_val.fill = copy(source_val.fill)
        target_val.alignment = copy(source_val.alignment)
        target_val.number_format = copy(source_val.number_format)
    
    target_val.value = lote.total_aportes
    ws.merge_cells(start_row=fila_total, start_column=6, end_row=fila_total, end_column=7)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f"Lote_{lote.numero_lote}_{lote.bodega_nombre}_{lote.fecha}.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

@login_required
def generar_pdf_lote(request, lote_id):
    lote = get_object_or_404(DepositoDiario, id=lote_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="resumen_lote_{lote.id}.pdf"'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - (1 * inch)
    x = 1 * inch
    
    p.setFont("Helvetica-Bold", 18)
    p.drawString(x, y, "Resumen de Cierre de Lote")
    y -= 30
    p.setFont("Helvetica", 12)
    p.drawString(x, y, f"Bodega: {lote.bodega_nombre}")
    y -= 20
    p.drawString(x, y, f"Fecha: {lote.fecha.strftime('%d-%m-%Y')}")
    y -= 20
    p.drawString(x, y, f"Lote: {lote.nombre_lote or 'Lote ' + str(lote.numero_lote)}")
    y -= 40
    p.setFont("Helvetica-Bold", 14)
    p.drawString(x, y, "Detalle de Aportes")
    y -= 25
    p.setFont("Helvetica", 10)
    for aporte in lote.aportes.all():
        monto_str = f"${aporte.monto:,}".replace(",", ".")
        p.drawString(x + 10, y, f"{aporte.trabajador.nombre}: {monto_str}")
        y -= 15
    p.setFont("Helvetica-Bold", 12)
    p.drawString(x, y - 10, f"Total Aportes: ${lote.total_aportes:,}".replace(",", "."))
    p.showPage()
    p.save()
    return response

# ====================================================================
#  NUEVA LÓGICA DE EXPORTACIÓN MENSUAL (SIMPLIFICADA - SOLO BILLETES)
# ====================================================================

@login_required
def exportar_mensual_excel(request):
    try:
        plantilla_path = os.path.join(settings.BASE_DIR, 'plantillas', 'PlantillaBancoMensual.xlsx')
        wb = openpyxl.load_workbook(plantilla_path)
        plantilla_sheet = wb.active 
    except FileNotFoundError:
        return HttpResponse("Error: PlantillaBancoMensual.xlsx no encontrada.", status=404)

    mes_sel = int(request.GET.get('mes', date.today().month))
    año_sel = int(request.GET.get('año', date.today().year))

    depositos_mes = DepositoDiario.objects.filter(
        fecha__year=año_sel,
        fecha__month=mes_sel
    ).order_by('fecha')

    dias_con_movimiento = sorted(list(set(d.fecha for d in depositos_mes)))
    meses_nombres = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses_nombres[mes_sel]

    # Mapas de filas para Billetes (Lado Derecho)
    mapa_mp = {20000: 8, 10000: 9, 5000: 10, 2000: 11, 1000: 12, 500: 13, 100: 14, 50: 15, 10: 16}
    mapa_dp = {20000: 23, 10000: 24, 5000: 25, 2000: 26, 1000: 27, 500: 28, 100: 29, 50: 30, 10: 31}
    mapa_tot = {20000: 8, 10000: 9, 5000: 10, 2000: 11, 1000: 12, 500: 13, 100: 14, 50: 15, 10: 16}

    for fecha in dias_con_movimiento:
        dia_num = str(fecha.day)
        ws = wb.copy_worksheet(plantilla_sheet)
        ws.title = dia_num
        
        # 1. Cabecera (General)
        ws['B2'] = nombre_mes
        ws['F2'] = fecha.day
        ws['I2'] = fecha.year
        ws['L2'] = fecha.strftime('%d-%m-%Y')

        # 2. Recopilar Datos (Billetes)
        lotes_dia = depositos_mes.filter(fecha=fecha)
        
        acum_mp = {k: {'cant': 0, 'monto': 0} for k in mapa_mp.keys()}
        acum_dp = {k: {'cant': 0, 'monto': 0} for k in mapa_dp.keys()}
        
        total_billetes_mp = 0
        total_billetes_dp = 0

        for lote in lotes_dia:
            es_mp = 'Manuel' in lote.bodega_nombre
            
            for des in lote.desglose.all():
                val = des.valor_unitario
                if es_mp:
                    if val in acum_mp:
                        acum_mp[val]['cant'] += des.cantidad
                        acum_mp[val]['monto'] += des.total_denominacion
                        total_billetes_mp += des.total_denominacion
                else:
                    if val in acum_dp:
                        acum_dp[val]['cant'] += des.cantidad
                        acum_dp[val]['monto'] += des.total_denominacion
                        total_billetes_dp += des.total_denominacion

        # 3. Escribir Tablas de Billetes (Lado Derecho)
        
        # Manuel Peñafiel (K8-M16)
        for val, info in acum_mp.items():
            fila = mapa_mp[val]
            ws[f'K{fila}'] = info['cant']
            ws[f'M{fila}'] = info['monto']
        ws['L18'] = total_billetes_mp # Total MP

        # David Perry (K23-M31)
        for val, info in acum_dp.items():
            fila = mapa_dp[val]
            ws[f'K{fila}'] = info['cant']
            ws[f'M{fila}'] = info['monto']
        ws['L33'] = total_billetes_dp # Total DP

        # Total del Día (Suma MP + DP) -> R8-T16
        total_gen_billetes = 0
        for val, fila in mapa_tot.items():
            cant_t = acum_mp[val]['cant'] + acum_dp[val]['cant']
            monto_t = acum_mp[val]['monto'] + acum_dp[val]['monto']
            ws[f'R{fila}'] = cant_t
            ws[f'T{fila}'] = monto_t
            total_gen_billetes += monto_t
        ws['S18'] = total_gen_billetes # Total General Día

    # Eliminar hoja plantilla
    if len(dias_con_movimiento) > 0:
        wb.remove(plantilla_sheet)
    else:
        plantilla_sheet.title = "Sin Movimientos"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f"{nombre_mes}-Depositos-{año_sel}.xlsx"
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    return response