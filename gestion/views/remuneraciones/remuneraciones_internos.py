import json
import os
import openpyxl
from io import BytesIO
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Prefetch, Sum
from django.utils import timezone

from gestion.models import (
    Trabajador, Remuneracion, RendicionDiaria, 
    AfpConfig, SaludConfig, AsignacionFamiliarConfig, 
    ConfiguracionGlobal, PlantillaLiquidacion
)
# Importamos las utilidades existentes
from .remuneraciones_utilidades import clean_currency, get_parametros_calculo

# ==========================================
# 1. NÓMINA Y NAVEGACIÓN
# ==========================================

def nomina_mensual(request):
    fecha_get = request.GET.get('fecha')
    hoy = timezone.now()
    
    if fecha_get:
        try:
            anio, mes = map(int, fecha_get.split('-'))
            fecha_seleccionada = date(anio, mes, 1)
        except ValueError:
            fecha_seleccionada = hoy.date()
    else:
        fecha_seleccionada = hoy.date()
    
    periodo_value = fecha_seleccionada.strftime('%Y-%m')
    periodo_str = fecha_seleccionada.strftime('%B %Y').title()

    trabajadores_qs = Trabajador.objects.filter(tipo='INTERNO', activo=True).order_by('nombre')
    liquidaciones_del_mes = Remuneracion.objects.filter(periodo=periodo_value)

    trabajadores_qs = trabajadores_qs.prefetch_related(
        Prefetch('remuneracion_set', queryset=liquidaciones_del_mes, to_attr='liquidacion_mes')
    )

    calculados_count = 0
    lista_trabajadores = []
    
    for t in trabajadores_qs:
        liq = t.liquidacion_mes[0] if t.liquidacion_mes else None
        if liq: calculados_count += 1
        lista_trabajadores.append({'obj': t, 'liquidacion': liq, 'tiene_calculo': bool(liq)})

    context = {
        'trabajadores_lista': lista_trabajadores,
        'periodo_str': periodo_str,
        'periodo_value': periodo_value,
        'total_count': trabajadores_qs.count(),
        'calculados_count': calculados_count,
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/nomina_mensual.html', context)

# ==========================================
# 2. CONFIGURACIÓN (PARÁMETROS)
# ==========================================

def parametros(request):
    afps = AfpConfig.objects.all().order_by('nombre')
    salud = SaludConfig.objects.all().order_by('nombre')
    tramos = AsignacionFamiliarConfig.objects.all().order_by('ingreso_tope')
    
    config_qs = ConfiguracionGlobal.objects.all()
    config_raw = {c.clave: c.valor for c in config_qs}

    if not tramos.exists():
        AsignacionFamiliarConfig.objects.create(tramo='A', ingreso_tope=0, monto_por_carga=0)
        tramos = AsignacionFamiliarConfig.objects.all().order_by('ingreso_tope')

    context = {
        'afps': afps, 
        'salud': salud, 
        'tramos': tramos,
        'config': {
            'gratificacion_legal_pct': config_raw.get('gratificacion_legal_pct', 0.0),
            'seguro_cesantia_pct': config_raw.get('seguro_cesantia_pct', 0.0),
        }
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_parametros/remuneraciones_parametros.html', context)

def actualizar_indicador_singular(request):
    if request.method == 'POST':
        clave = request.POST.get('clave') 
        valor = request.POST.get('valor')
        descripcion = request.POST.get('descripcion', 'Indicador')
        if clave and valor:
            try:
                ConfiguracionGlobal.objects.update_or_create(
                    clave=clave,
                    defaults={'valor': float(valor.replace(',', '.')), 'descripcion': descripcion}
                )
            except: pass
    return redirect('parametros_remuneraciones')

def crear_entidad_previsional(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo_entidad')
        nombre = request.POST.get('nombre')
        tasa = request.POST.get('tasa')
        if nombre and tasa:
            try:
                tasa_val = float(tasa.replace(',', '.'))
                if tipo == 'afp': AfpConfig.objects.create(nombre=nombre, tasa=tasa_val)
                elif tipo == 'salud': SaludConfig.objects.create(nombre=nombre, tasa=tasa_val)
            except: pass 
    return redirect('parametros_remuneraciones')

def editar_entidad_previsional(request, tipo, id):
    if request.method == 'POST':
        klass = AfpConfig if tipo == 'afp' else SaludConfig
        instance = get_object_or_404(klass, id=id)
        try:
            instance.nombre = request.POST.get('nombre')
            instance.tasa = float(request.POST.get('tasa').replace(',', '.'))
            instance.save()
        except: pass
    return redirect('parametros_remuneraciones')

def eliminar_entidad_previsional(request, tipo, id):
    if request.method == 'POST':
        klass = AfpConfig if tipo == 'afp' else SaludConfig
        get_object_or_404(klass, id=id).delete()
    return redirect('parametros_remuneraciones')

def editar_tramos_asignacion(request):
    if request.method == 'POST':
        for tramo in AsignacionFamiliarConfig.objects.all():
            try:
                tramo.ingreso_tope = int(clean_currency(request.POST.get(f'tope_{tramo.tramo}', '0')))
                tramo.monto_por_carga = int(clean_currency(request.POST.get(f'monto_{tramo.tramo}', '0')))
                tramo.save()
            except: pass
    return redirect('parametros_remuneraciones')

# ==========================================
# 3. CÁLCULO DE SUELDO
# ==========================================

def calcular_sueldo(request, id):
    try:
        liq_ex = Remuneracion.objects.select_related('trabajador').get(id=id)
        trabajador = liq_ex.trabajador
        periodo_actual = liq_ex.periodo
    except Remuneracion.DoesNotExist:
        trabajador = get_object_or_404(Trabajador, id=id)
        liq_ex = None
        periodo_actual = request.GET.get('periodo') or date.today().strftime('%Y-%m')

    p = get_parametros_calculo()
    
    if request.method == 'POST':
        dias = int(request.POST.get('dias') or 30)
        horas_extras = float(request.POST.get('horas_extras') or 0)
        anticipo = clean_currency(request.POST.get('anticipo', 0))
        abono_faltante = clean_currency(request.POST.get('abono_faltante', 0))

        # Listas JSON
        l_haberes = json.loads(request.POST.get('detalle_haberes', '[]'))
        l_asignaciones = json.loads(request.POST.get('detalle_asignaciones', '[]'))
        l_bonos = json.loads(request.POST.get('detalle_bonos', '[]'))
        l_descuentos = json.loads(request.POST.get('detalle_descuentos', '[]'))
        l_legales_raw = json.loads(request.POST.get('detalle_descuentos_legales', '[]'))

        # Sumas manuales
        tot_hab_imp = sum(int(x.get('monto',0)) for x in l_haberes)
        tot_asig = sum(int(x.get('monto',0)) for x in l_asignaciones)
        tot_no_imp = sum(int(x.get('monto',0)) for x in l_bonos)
        tot_otros_desc = sum(int(x.get('monto',0)) for x in l_descuentos)

        # Cálculos Base
        sueldo_base_prop = (Decimal(trabajador.sueldo_base) / 30 * Decimal(dias)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        monto_he = (Decimal(horas_extras) * Decimal(trabajador.valor_hora_extra)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        base_pre_grat = sueldo_base_prop + monto_he + Decimal(tot_hab_imp)
        grat_bruta = min(base_pre_grat * p['gratificacion_pct'], p.get('tope_gratificacion', 0)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        sueldo_imponible = base_pre_grat + grat_bruta
        base_desc = min(sueldo_imponible, p['tope_imponible'])

        # Descuentos Legales Variables
        l_legales_proc = []
        tot_legales_extra = Decimal(0)
        for item in l_legales_raw:
            try:
                val = Decimal(str(item.get('valor', 0)))
                monto = (sueldo_imponible * (val/100)).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if item.get('tipo') == '%' else val
                tot_legales_extra += monto
                l_legales_proc.append({**item, 'monto_calculado': int(monto)})
            except: continue

        # Legales Estándar
        m_afp = (base_desc * Decimal(trabajador.afp.tasa)/100).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if trabajador.afp else Decimal(0)
        m_salud = (base_desc * Decimal(trabajador.salud.tasa)/100).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if trabajador.salud else Decimal(0)
        
        tasa_cesantia = p['seguro_cesantia_pct']
        if trabajador.fecha_ingreso:
            antiguedad = (date.today() - trabajador.fecha_ingreso).days / 365.25
            if antiguedad >= 11: tasa_cesantia = Decimal(0)
        
        m_cesantia = (base_desc * tasa_cesantia).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

        # Asignación Familiar
        m_af_auto = Decimal(0)
        tramo_letra = None
        if trabajador.tiene_asignacion_familiar and trabajador.cargas_familiares > 0:
            for tr in p['tramos_af']:
                if sueldo_imponible <= Decimal(tr.ingreso_tope):
                    tramo_letra = tr.tramo
                    m_af_auto = Decimal(tr.monto_por_carga) * Decimal(trabajador.cargas_familiares)
                    break
        
        # Impuesto
        afecto = sueldo_imponible - (m_afp + m_salud + m_cesantia)
        m_impuesto = max(Decimal(0), (afecto * Decimal('0.04')) - Decimal('37218.42')).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if afecto > 0 else Decimal(0)

        # Faltante Caja
        m_faltante_auto = Decimal(0)
        if trabajador.filtro_trabajador:
            y, m = map(int, periodo_actual.split('-'))
            bal = RendicionDiaria.objects.filter(trabajador=trabajador, fecha__year=y, fecha__month=m).aggregate(Sum('diferencia'))['diferencia__sum'] or 0
            if bal < 0: m_faltante_auto = Decimal(abs(bal))
        
        m_faltante_fin = max(Decimal(0), m_faltante_auto - Decimal(abono_faltante))

        # Totales Finales
        tot_haberes = sueldo_imponible + m_af_auto + Decimal(tot_asig) + Decimal(tot_no_imp)
        tot_desc = m_afp + m_salud + m_cesantia + m_impuesto + tot_legales_extra + Decimal(anticipo) + Decimal(tot_otros_desc) + m_faltante_fin
        liquido = tot_haberes - tot_desc

        detalle = {
            'sueldo_base_proporcional': str(sueldo_base_prop),
            'dias_trabajados': dias,
            'horas_extras': horas_extras,
            'monto_horas_extras': str(monto_he),
            'lista_haberes': l_haberes,
            'lista_asignaciones': l_asignaciones,
            'lista_bonos': l_bonos,
            'lista_descuentos': l_descuentos,
            'lista_descuentos_legales': l_legales_proc,
            'monto_anticipo': str(anticipo),
            'faltante_automatico_original': str(m_faltante_auto),
            'abono_faltante': str(abono_faltante),
            'asignacion_familiar': str(m_af_auto),
            'sueldo_imponible': str(sueldo_imponible),
            'gratificacion': str(grat_bruta),
            'afp_tasa': float(trabajador.afp.tasa) if trabajador.afp else 0,
            'afp_nombre': trabajador.afp.nombre if trabajador.afp else "Sin AFP",
            'monto_afp': str(m_afp),
            'salud_tasa': float(trabajador.salud.tasa) if trabajador.salud else 0,
            'salud_nombre': trabajador.salud.nombre if trabajador.salud else "Sin Salud",
            'monto_salud': str(m_salud),
            'monto_seguro_cesantia': str(m_cesantia),
            'tramo_letra': tramo_letra,
            'monto_impuesto': str(m_impuesto),
            'afecto_impuesto': str(max(Decimal(0), afecto)),
            'monto_faltante': str(m_faltante_fin),
            'periodo': periodo_actual,
            'cumplio_11_anos': bool(tasa_cesantia == 0 and trabajador.fecha_ingreso)
        }

        defaults = {
            'sueldo_liquido': int(liquido), 'total_haberes': int(tot_haberes), 'total_descuentos': int(tot_desc),
            'monto_impuesto': int(m_impuesto), 'monto_faltante': int(m_faltante_fin), 'detalle_json': detalle
        }

        if liq_ex:
            for k, v in defaults.items(): setattr(liq_ex, k, v)
            liq_ex.save()
            id_fin = liq_ex.id
        else:
            id_fin = Remuneracion.objects.create(trabajador=trabajador, periodo=periodo_actual, **defaults).id
        
        return redirect('ver_liquidacion', id=id_fin)

    # GET REQUEST
    initial = {}
    faltante_auto = 0
    if liq_ex:
        d = liq_ex.detalle_json
        initial = {
            'dias': d.get('dias_trabajados', 30), 'horas_extras': d.get('horas_extras', 0),
            'anticipo': int(float(d.get('monto_anticipo', 0))), 'abono_faltante': int(float(d.get('abono_faltante', 0))),
            'json_haberes': json.dumps(d.get('lista_haberes', [])), 'json_asignaciones': json.dumps(d.get('lista_asignaciones', [])),
            'json_bonos': json.dumps(d.get('lista_bonos', [])), 'json_descuentos': json.dumps(d.get('lista_descuentos', [])),
            'json_descuentos_legales': json.dumps(d.get('lista_descuentos_legales', []))
        }
        faltante_auto = int(float(d.get('faltante_automatico_original', 0)))
    else:
        # Calcular faltante inicial
        if trabajador.filtro_trabajador:
            y, m = map(int, periodo_actual.split('-'))
            bal = RendicionDiaria.objects.filter(trabajador=trabajador, fecha__year=y, fecha__month=m).aggregate(Sum('diferencia'))['diferencia__sum'] or 0
            if bal < 0: faltante_auto = abs(bal)
        initial = {'dias': 30, 'horas_extras': 0, 'anticipo': 0, 'abono_faltante': 0}

    context = {
        'trabajador': trabajador,
        'periodo_str': date(int(periodo_actual[:4]), int(periodo_actual[5:]), 1).strftime('%B %Y').upper(),
        'initial': initial,
        'faltante_automatico': faltante_auto,
        'tope_grat_anual': p.get('tope_gratificacion', 0),
        'periodo_mes_nombre': date(int(periodo_actual[:4]), int(periodo_actual[5:]), 1).strftime('%B')
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/formulario_calculo.html', context)

def ver_liquidacion(request, id):
    liquidacion = get_object_or_404(Remuneracion.objects.select_related('trabajador'), id=id)
    detalle = liquidacion.detalle_json
    
    det_fmt = {}
    for k, v in detalle.items():
        try: 
            # Intentar convertir montos a int para quitar decimales visuales, excepto tasas y horas
            if k in ['afp_tasa', 'salud_tasa', 'horas_extras']:
                det_fmt[k] = float(v)
            else:
                if isinstance(v, (int, float, str)) and str(v).replace('.','',1).isdigit():
                    det_fmt[k] = int(float(v))
                else:
                    det_fmt[k] = v
        except: det_fmt[k] = v

    # Rut y Bodega
    rut = liquidacion.trabajador.rut
    bod = liquidacion.trabajador.bodega_asignada
    if bod == 'Ambos': bod = liquidacion.trabajador.bodega_facturacion
    bod_nom = "Manuel Peñafiel" if bod == '1221' else ("David Perry" if bod == '1225' else "Desconocida")

    context = {
        'liquidacion': liquidacion, 't': liquidacion.trabajador, 'rut_fmt': rut,
        'bodega_real': bod_nom, 'detalle': det_fmt, 'total_haberes_brutos': liquidacion.total_haberes,
        'periodo_str': date(int(det_fmt['periodo'][:4]), int(det_fmt['periodo'][5:]), 1).strftime('%B %Y').upper(),
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/liquidacion_detalle.html', context)

# ======================================================
# 4. EXPORTACIÓN EXCEL
# ======================================================

def _get_plantilla_path():
    """Retorna la ruta de la plantilla Excel actualizada"""
    # CORRECCIÓN: La carpeta 'plantillas' está en la raíz del proyecto
    return os.path.join(settings.BASE_DIR, 'plantillas', 'PagoInternosPlantilla.xlsx')

def _obtener_direccion_pago(trabajador):
    """Determina la dirección basada en la bodega de facturación o asignada"""
    # Prioridad: Bodega Facturación -> Bodega Asignada
    bodega = trabajador.bodega_facturacion if trabajador.bodega_facturacion else trabajador.bodega_asignada
    
    # Normalizamos a string por si es None
    bodega = str(bodega)
    
    if '1225' in bodega or 'David' in bodega or 'Perry' in bodega:
        return "Avda. David Perry N° 647"
    elif '1221' in bodega or 'Manuel' in bodega or 'Peñafiel' in bodega:
        return "Avda. Manuel Peñafiel N° 1295"
    else:
        return "Avda. Manuel Peñafiel N° 1295"

def _llenar_hoja_liquidacion(ws, liquidacion):
    """
    Rellena la hoja siguiendo las instrucciones estrictas de celdas.
    """
    t = liquidacion.trabajador
    d = liquidacion.detalle_json 
    
    # Parsear fecha del periodo (YYYY-MM)
    anio_str, mes_str = liquidacion.periodo.split('-')
    fecha_dt = date(int(anio_str), int(mes_str), 1)
    
    MESES = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    nombre_mes = MESES[fecha_dt.month] 
    anio = fecha_dt.year               

    # 1. El nombre de la hoja se gestiona externamente (primer nombre)

    # 2. I2: Mes (Capitalizado)
    ws['I2'] = nombre_mes

    # 3. J2: Año
    ws['J2'] = anio

    # 4. I4: Sueldo Líquido Real
    ws['I4'] = liquidacion.sueldo_liquido

    # 5. E8: Nombre Completo (Capitalizado Ej: Leonel Araya Julio)
    ws['E8'] = t.nombre.title()

    # 6. E9: Cargo Completo
    ws['E9'] = t.cargo or "Operario"

    # 7. E10: Dirección según bodega
    ws['E10'] = _obtener_direccion_pago(t)

    # 8. E11: Días trabajados (recuperado del JSON detalle)
    ws['E11'] = d.get('dias_trabajados', 30)

    # 9. E12: Gratificación estándar (0.25 y formato Porcentaje)
    ws['E12'] = 0.25
    ws['E12'].number_format = '0%'

    # 10. I8: RUT con formato (12.426.714-5)
    rut_limpio = str(t.rut).replace('.', '') # Aseguramos limpiar primero
    if '-' in rut_limpio:
        parte_num, dv = rut_limpio.split('-')
        try:
            # Formatear la parte numérica con separador de miles y reemplazar comas por puntos
            parte_num_fmt = "{:,}".format(int(parte_num)).replace(",", ".")
            ws['I8'] = f"{parte_num_fmt}-{dv}"
        except ValueError:
            ws['I8'] = t.rut
    else:
        ws['I8'] = t.rut

    # 11. I10: Fecha Ingreso (DD-MM-YYYY)
    if t.fecha_ingreso:
        ws['I10'] = t.fecha_ingreso.strftime('%d-%m-%Y')
    else:
        ws['I10'] = ""

    # 12. I11: Sueldo Diario (Base / 30) - CON PUNTO Y FORMATO NÚMERO
    if t.sueldo_base:
        sueldo_diario = int(t.sueldo_base / 30)
        ws['I11'] = sueldo_diario
        ws['I11'].number_format = '#,##0' # Formato Excel para separador de miles
    else:
        ws['I11'] = 0

    # REQUERIMIENTOS ANTERIORES:
    
    # E16: Sueldo Base Original
    ws['E16'] = t.sueldo_base

    # E17: Horas Extras (Cantidad)
    try:
        horas = float(d.get('horas_extras', 0))
        # Si es entero visualmente (ej: 5.0), lo mostramos como 5
        ws['E17'] = int(horas) if horas.is_integer() else horas
    except:
        ws['E17'] = 0

    # I16: Sueldo Real (Proporcional)
    try:
        s_real = float(d.get('sueldo_base_proporcional', 0))
        ws['I16'] = int(s_real)
    except:
        s_real = 0
        ws['I16'] = 0

    # I17: Ganancia Horas Extras
    try:
        ganancia_he = float(d.get('monto_horas_extras', 0))
        ws['I17'] = int(ganancia_he)
    except:
        ganancia_he = 0
        ws['I17'] = 0

    # I18: Gratificación (Valor calculado)
    try:
        grat = float(d.get('gratificacion', 0))
        ws['I18'] = int(grat)
    except:
        grat = 0
        ws['I18'] = 0

    # I19: Subtotal (Suma de Real + Ganancia HE + Gratificación)
    subtotal = int(s_real) + int(ganancia_he) + int(grat)
    ws['I19'] = subtotal

    # ===============================================
    # OTROS HABERES (Bonos no imponibles)
    # ===============================================
    # Ubicación: C21 (Texto), I21 (Monto). Incrementa fila.
    # Subtotal en I24.
    
    lista_bonos = d.get('lista_bonos', [])
    fila_actual = 21
    total_otros_haberes = 0
    
    for bono in lista_bonos:
        if fila_actual >= 24: break 
        monto = int(float(bono.get('monto', 0)))
        desc = bono.get('desc', '')
        ws[f'C{fila_actual}'] = desc
        ws[f'I{fila_actual}'] = monto
        total_otros_haberes += monto
        fila_actual += 1
        
    ws['I24'] = total_otros_haberes 

    # ===============================================
    # ASIGNACIONES FAMILIARES
    # ===============================================
    # Ubicación: C26 (Texto), I26 (Monto). Incrementa fila.
    # Subtotal en I28.
    
    fila_actual_asig = 26
    total_asignaciones = 0
    
    asig_auto = int(float(d.get('asignacion_familiar', 0)))
    if asig_auto > 0:
        ws[f'C{fila_actual_asig}'] = f"Asig. Familiar ({t.cargas_familiares} Cargas)"
        ws[f'I{fila_actual_asig}'] = asig_auto
        total_asignaciones += asig_auto
        fila_actual_asig += 1
        
    lista_asig_manual = d.get('lista_asignaciones', [])
    for item in lista_asig_manual:
        if fila_actual_asig >= 28: break 
        monto = int(float(item.get('monto', 0)))
        desc = item.get('desc', '')
        ws[f'C{fila_actual_asig}'] = desc
        ws[f'I{fila_actual_asig}'] = monto
        total_asignaciones += monto
        fila_actual_asig += 1
        
    ws['I28'] = total_asignaciones

    # ===============================================
    # DESCUENTOS LEGALES (AFP, SALUD, CESANTÍA)
    # ===============================================
    
    # 1. AFP
    try:
        tasa_afp = float(d.get('afp_tasa', 0)) / 100
        monto_afp = int(float(d.get('monto_afp', 0)))
        nombre_afp = d.get('afp_nombre', '')
    except:
        tasa_afp, monto_afp, nombre_afp = 0, 0, ''

    ws['C30'] = tasa_afp
    ws['C30'].number_format = '0.00%'
    ws['E30'] = f"Cot. A.F.P. {nombre_afp}"
    ws['J30'] = monto_afp

    # 2. SALUD
    try:
        tasa_salud = float(d.get('salud_tasa', 0)) / 100
        monto_salud = int(float(d.get('monto_salud', 0)))
        nombre_salud = d.get('salud_nombre', 'Fonasa')
    except:
        tasa_salud, monto_salud, nombre_salud = 0, 0, ''

    ws['C31'] = tasa_salud
    ws['C31'].number_format = '0.00%'
    ws['E31'] = nombre_salud
    ws['J31'] = monto_salud

    # 3. SEGURO CESANTÍA
    try:
        monto_cesantia = int(float(d.get('monto_seguro_cesantia', 0)))
        cumplio_11 = d.get('cumplio_11_anos', False)
        
        # Lógica de texto y porcentaje según reglas
        if cumplio_11:
            texto_cesantia = "Seguro de Cesantía ( cumplió 11 años )"
            tasa_cesantia = 0
        else:
            texto_cesantia = "Seguro de Cesantía"
            # Si hay monto, asumimos 0.6% (aprox), si es 0, es 0%
            tasa_cesantia = 0.006 if monto_cesantia > 0 else 0
            
    except:
        monto_cesantia, tasa_cesantia = 0, 0
        texto_cesantia = "Seguro de Cesantía"

    ws['C32'] = tasa_cesantia
    ws['C32'].number_format = '0.0%'
    ws['E32'] = texto_cesantia
    ws['J32'] = monto_cesantia

    # 4. SUBTOTAL LEGALES (J33)
    ws['J33'] = monto_afp + monto_salud + monto_cesantia

    # ===============================================
    # OTRA SECCIÓN (ANTICIPOS Y FALTANTES)
    # ===============================================

    # 1. ANTICIPO (J42)
    try:
        anticipo = int(float(d.get('monto_anticipo', 0)))
    except:
        anticipo = 0
    ws['J42'] = anticipo

    # 2. FALTANTE (C43, J43)
    try:
        faltante = int(float(d.get('monto_faltante', 0)))
    except:
        faltante = 0

    if faltante > 0:
        ws['C43'] = "Faltante en Caja"
        ws['J43'] = faltante
    else:
        # Limpiar si no hay faltante (por si la plantilla tuviera algo)
        ws['C43'] = "" 
        ws['J43'] = 0

    # 3. SUBTOTAL SECCIÓN (J44)
    ws['J44'] = anticipo + faltante

    # ===============================================
    # ÚLTIMA SECCIÓN (TOTALES Y FIRMA)
    # ===============================================

    # I46: Total Haberes
    ws['I46'] = liquidacion.total_haberes
    
    # J46: Total Descuentos
    ws['J46'] = liquidacion.total_descuentos
    
    # J47: Sueldo Líquido
    ws['J47'] = liquidacion.sueldo_liquido
    
    # C52: Nombre Completo nuevamente
    ws['C52'] = t.nombre.title()
    
    # J50: Sueldo Líquido nuevamente
    ws['J50'] = liquidacion.sueldo_liquido

def exportar_liquidacion_excel(request, id):
    """Genera Excel Individual: [PrimerNombre] [Mes] [Año].xlsx"""
    remu = get_object_or_404(Remuneracion, id=id)
    ruta = _get_plantilla_path()
    
    if not os.path.exists(ruta):
        return HttpResponse(f"Error: No se encontró la plantilla en {ruta}", status=404)

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    
    # Nombre de la hoja: Primer nombre del trabajador
    primer_nombre = remu.trabajador.nombre.strip().split()[0]
    ws.title = primer_nombre

    _llenar_hoja_liquidacion(ws, remu)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Nombre archivo: Cristóbal Julio 2025.xlsx
    anio_str, mes_str = remu.periodo.split('-')
    MESES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
             7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
    mes_nombre = MESES[int(mes_str)]
    
    filename = f"{primer_nombre} {mes_nombre} {anio_str}.xlsx"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def exportar_excel_global(request):
    """Genera Excel Masivo: Liquidaciones [Mes] [Año] OPL.xlsx"""
    fecha_get = request.GET.get('fecha') # Formato YYYY-MM
    if not fecha_get: return HttpResponse("Falta fecha", status=400)

    remuneraciones = Remuneracion.objects.filter(periodo=fecha_get).select_related('trabajador').order_by('trabajador__nombre')
    if not remuneraciones.exists():
        return HttpResponse("No hay datos para este periodo", status=404)

    ruta = _get_plantilla_path()
    if not os.path.exists(ruta):
        return HttpResponse(f"Error: Plantilla no encontrada en {ruta}", status=404)

    # Cargar base
    wb_base = openpyxl.load_workbook(ruta)
    ws_template = wb_base.active 
    ws_template.title = "Template"

    # Variables para nombre del archivo final
    anio_str, mes_str = fecha_get.split('-')
    MESES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
             7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
    mes_nombre = MESES[int(mes_str)]

    # Diccionario para controlar nombres de hojas repetidos
    nombres_usados = {}

    for remu in remuneraciones:
        # Clonar hoja template
        target_sheet = wb_base.copy_worksheet(ws_template)
        
        # Definir nombre de hoja: Primer nombre
        primer_nombre = remu.trabajador.nombre.strip().split()[0]
        
        # Evitar duplicados
        if primer_nombre in nombres_usados:
            nombres_usados[primer_nombre] += 1
            nombre_hoja = f"{primer_nombre} {nombres_usados[primer_nombre]}"
        else:
            nombres_usados[primer_nombre] = 1
            nombre_hoja = primer_nombre
            
        target_sheet.title = nombre_hoja
        
        _llenar_hoja_liquidacion(target_sheet, remu)

    # Eliminar la plantilla original
    wb_base.remove(ws_template)

    buffer = BytesIO()
    wb_base.save(buffer)
    buffer.seek(0)
    
    # Nombre archivo: Liquidaciones Julio 2025 OPL.xlsx
    filename = f"Liquidaciones {mes_nombre} {anio_str} OPL.xlsx"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response