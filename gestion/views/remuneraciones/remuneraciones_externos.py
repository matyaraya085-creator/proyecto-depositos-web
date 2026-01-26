import json
import calendar
import os
import openpyxl
from copy import copy  # <--- IMPORTANTE: Necesario para copiar estilos
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from gestion.models import Trabajador, RemuneracionExterna, RendicionDiaria, TarifaComision

# Helper para nombres en español
def get_nombre_mes(mes_num):
    meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    return meses.get(mes_num, "")

def nomina_externos(request):
    fecha_get = request.GET.get('fecha')
    fecha_dt = datetime.strptime(fecha_get, '%Y-%m') if fecha_get else timezone.now()
    mes, anio = fecha_dt.month, fecha_dt.year

    trabajadores = Trabajador.objects.filter(tipo='EXTERNO', activo=True).order_by('nombre')
    procesados = RemuneracionExterna.objects.filter(mes=mes, anio=anio).values_list('trabajador_id', flat=True)

    context = {
        'trabajadores': trabajadores, 'mes': mes, 'anio': anio, 'procesados': procesados,
        'fecha_actual': fecha_dt.strftime('%Y-%m'), 'mes_nombre': calendar.month_name[mes].capitalize()
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_externos/sueldos_externos.html', context)

def calcular_remuneracion_externa(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    mes = int(request.GET.get('mes', timezone.now().month))
    anio = int(request.GET.get('anio', timezone.now().year))

    tarifas = TarifaComision.objects.last() or TarifaComision()
    
    # 1. Obtener datos automáticos desde Rendiciones
    rendiciones = RendicionDiaria.objects.filter(trabajador=trabajador, fecha__month=mes, fecha__year=anio)

    conteo = {'5kg': 0, '11kg': 0, '15kg': 0, '45kg': 0, 'cat5kg': 0, 'cat15kg': 0, 'ultra15': 0}
    faltante_acum_db = 0
    anticipo_acum_db = 0

    for r in rendiciones:
        conteo['5kg'] += (r.gas_5kg or 0)
        conteo['11kg'] += (r.gas_11kg or 0)
        conteo['15kg'] += (r.gas_15kg or 0)
        conteo['45kg'] += (r.gas_45kg or 0)
        conteo['cat5kg'] += (r.gasc_5kg or 0)
        conteo['cat15kg'] += (r.gasc_15kg or 0)
        conteo['ultra15'] += (r.gas_ultra_15kg or 0)
        
        anticipo_acum_db += (r.monto_anticipo or 0)
        if r.diferencia < 0: 
            faltante_acum_db += abs(r.diferencia)

    pago_cilindros = (
        conteo['5kg']*tarifas.tarifa_5kg + conteo['11kg']*tarifas.tarifa_11kg + 
        conteo['15kg']*tarifas.tarifa_15kg + conteo['45kg']*tarifas.tarifa_45kg +
        conteo['cat5kg']*tarifas.tarifa_cat_5kg + conteo['cat15kg']*tarifas.tarifa_cat_15kg + 
        conteo['ultra15']*tarifas.tarifa_ultra_15kg
    )

    remu_existente = RemuneracionExterna.objects.filter(trabajador=trabajador, mes=mes, anio=anio).first()

    if request.method == 'POST':
        # --- A. INGRESOS ---
        asistencia = int(request.POST.get('asistencia_tecnica') or 0)
        subtotal_neto = pago_cilindros + asistencia
        iva = int(subtotal_neto * 0.19)
        total_ingresos = subtotal_neto + iva

        # --- B. DESCUENTOS ---
        anticipo_extra = int(request.POST.get('anticipo_extra') or 0)
        total_anticipo = anticipo_acum_db + anticipo_extra

        faltante_extra = int(request.POST.get('faltante_extra') or 0)
        total_faltante = faltante_acum_db + faltante_extra

        # Lista dinámica
        json_otros = request.POST.get('json_otros_descuentos', '[]')
        try:
            lista_otros = json.loads(json_otros)
            total_otros = sum(int(item.get('monto', 0)) for item in lista_otros)
        except:
            lista_otros = []
            total_otros = 0

        # SUMA FINAL DESCUENTOS (Sin préstamo explícito en model, usamos total_otros como préstamo/varios)
        total_descuentos = total_anticipo + total_faltante + total_otros

        # --- C. FINAL ---
        monto_final = total_ingresos - total_descuentos

        remu, _ = RemuneracionExterna.objects.update_or_create(
            trabajador=trabajador, mes=mes, anio=anio,
            defaults={
                'nro_factura': request.POST.get('nro_factura'),
                'pago_cilindros': pago_cilindros,
                'asistencia_tecnica': asistencia,
                'subtotal_neto': subtotal_neto,
                'iva': iva,
                'total_bruto': total_ingresos,
                'anticipo_base': anticipo_acum_db,
                'anticipo_extra': anticipo_extra,
                'faltante_base': faltante_acum_db,
                'faltante_extra': faltante_extra,
                'json_otros_descuentos': json_otros,
                'total_otros_descuentos': total_otros,
                'total_descuentos': total_descuentos,
                'monto_total_pagar': monto_final,
                'json_detalle_cilindros': json.dumps(conteo)
            }
        )
        return redirect('detalle_remuneracion_externa', remu_id=remu.id)

    return render(request, 'gestion/remuneraciones/remuneraciones_externos/calculo_externos.html', {
        'trabajador': trabajador, 
        'mes': mes, 
        'anio': anio,
        'conteo': conteo,
        'pago_cilindros': pago_cilindros,
        'anticipo_db': anticipo_acum_db,
        'faltante_db': faltante_acum_db,
        'remu': remu_existente
    })

def detalle_remuneracion_externa(request, remu_id):
    remu = get_object_or_404(RemuneracionExterna, id=remu_id)
    
    try: detalle_otros = json.loads(remu.json_otros_descuentos)
    except: detalle_otros = []

    try: detalle_cilindros = json.loads(remu.json_detalle_cilindros)
    except: detalle_cilindros = {}

    return render(request, 'gestion/remuneraciones/remuneraciones_externos/detalle_externos.html', {
        'remu': remu,
        'detalle_otros': detalle_otros,
        'detalle_cilindros': detalle_cilindros
    })

def exportar_excel_externos(request):
    # 1. Obtener Fechas
    fecha_get = request.GET.get('fecha')
    if fecha_get:
        fecha_dt = datetime.strptime(fecha_get, '%Y-%m')
        mes, anio = fecha_dt.month, fecha_dt.year
    else:
        hoy = timezone.now()
        mes, anio = hoy.month, hoy.year
    
    nombre_mes = get_nombre_mes(mes)

    # 2. Cargar Plantilla
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'plantillas', 'PagoFleterosPlantilla.xlsx')
    
    if not os.path.exists(ruta_plantilla):
        return HttpResponse(f"Error: No se encontró la plantilla en {ruta_plantilla}", status=404)

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active # Hoja activa
    
    # 3. Configurar Encabezados
    ws.title = f"{nombre_mes} {anio}"
    ws['C2'] = f"Pago Fletes mes de {nombre_mes} {anio} - Servicios Logísticos Celestina Araya SpA"

    # 4. Obtener Datos
    remuneraciones = RemuneracionExterna.objects.filter(mes=mes, anio=anio).select_related('trabajador').order_by('trabajador__nombre')

    if not remuneraciones.exists():
        return HttpResponse("No hay datos calculados para este mes.", status=404)

    # 5. Insertar y Estilizar Filas
    fila_inicio = 6
    cantidad_trabajadores = remuneraciones.count()

    # Si hay más de 1 trabajador, insertamos filas nuevas
    if cantidad_trabajadores > 1:
        # insert_rows(indice_donde_insertar, cantidad)
        # Se insertan DESPUÉS de la fila 6 original
        ws.insert_rows(fila_inicio + 1, amount=cantidad_trabajadores - 1)
        
        # --- NUEVO: COPIAR ESTILOS DE LA FILA 6 A LAS NUEVAS ---
        # Recorremos desde la fila 7 hasta la última nueva insertada
        for row_idx in range(fila_inicio + 1, fila_inicio + cantidad_trabajadores):
            for col_idx in range(1, ws.max_column + 1):
                # Celda origen (Plantilla Fila 6)
                source_cell = ws.cell(row=fila_inicio, column=col_idx)
                # Celda destino (Nueva fila)
                target_cell = ws.cell(row=row_idx, column=col_idx)
                
                # Copiar atributos de estilo
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

    # 6. Rellenar Datos
    for i, remu in enumerate(remuneraciones):
        fila_actual = fila_inicio + i
        
        # B: Número correlativo
        ws[f'B{fila_actual}'] = i + 1
        
        # C: Nombre Trabajador
        ws[f'C{fila_actual}'] = remu.trabajador.nombre
        
        # D: Nro Factura
        ws[f'D{fila_actual}'] = remu.nro_factura if remu.nro_factura else "S/N"
        
        # E: Neto (Subtotal Neto)
        ws[f'E{fila_actual}'] = remu.subtotal_neto
        
        # F: IVA
        ws[f'F{fila_actual}'] = remu.iva
        
        # G: Total (Bruto)
        ws[f'G{fila_actual}'] = remu.total_bruto
        
        # H: Préstamo (Mapeado a Total Otros Descuentos)
        ws[f'H{fila_actual}'] = remu.total_otros_descuentos
        
        # I: Anticipo (Base + Extra)
        total_anticipo = (remu.anticipo_base or 0) + (remu.anticipo_extra or 0)
        ws[f'I{fila_actual}'] = total_anticipo
        
        # J: Faltante (Base + Extra)
        total_faltante = (remu.faltante_base or 0) + (remu.faltante_extra or 0)
        ws[f'J{fila_actual}'] = total_faltante
        
        # K: Líquido a Pago
        ws[f'K{fila_actual}'] = remu.monto_total_pagar

    # 7. Guardar y Responder
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"Pago Fleteros OPL {nombre_mes} {anio}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response