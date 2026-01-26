import re
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Prefetch, Sum
from django.utils import timezone
from django.http import HttpResponse
from gestion.models import AfpConfig, SaludConfig, AsignacionFamiliarConfig, ConfiguracionGlobal, Trabajador, Remuneracion, RendicionDiaria, PlantillaLiquidacion, TarifaComision, RemuneracionExterna
from django.db import IntegrityError
from datetime import date, datetime
import locale
import calendar
from decimal import Decimal, ROUND_HALF_UP
import json
import openpyxl
from openpyxl.styles import Font, Alignment

# Intentamos configurar locale a español para los nombres de meses
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass

# ==========================================
# 1. VISTAS DE NAVEGACIÓN
# ==========================================

def menu_remuneraciones(request):
    """
    Vista que renderiza el menú principal de remuneraciones.
    """
    return render(request, 'gestion/remuneraciones/remuneraciones_menu.html')

def nomina_mensual(request):
    """
    Muestra la nómina filtrando trabajadores INTERNOS y ACTIVOS.
    """
    # 1. Gestión de Fechas
    fecha_get = request.GET.get('fecha') # Formato YYYY-MM
    hoy = timezone.now()
    
    if fecha_get:
        try:
            anio = int(fecha_get.split('-')[0])
            mes = int(fecha_get.split('-')[1])
            fecha_seleccionada = date(anio, mes, 1)
        except ValueError:
            fecha_seleccionada = hoy.date()
    else:
        fecha_seleccionada = hoy.date()
    
    # String para el input type="month" (YYYY-MM)
    periodo_value = fecha_seleccionada.strftime('%Y-%m')
    # String para mostrar al usuario (Ej: Enero 2026)
    periodo_str = fecha_seleccionada.strftime('%B %Y').title()

    # 2. Filtro de Trabajadores
    # - Tipo: INTERNO (Contrato)
    # - Activo: True (Trabajando actualmente)
    trabajadores_qs = Trabajador.objects.filter(
        tipo='INTERNO', 
        activo=True
    ).order_by('nombre')

    # 3. Optimización con Prefetch
    # Buscamos solo las liquidaciones que coincidan EXACTAMENTE con el periodo seleccionado
    liquidaciones_del_mes = Remuneracion.objects.filter(periodo=periodo_value)

    # Adjuntamos la liquidación al objeto trabajador si existe
    trabajadores_qs = trabajadores_qs.prefetch_related(
        Prefetch('remuneracion_set', queryset=liquidaciones_del_mes, to_attr='liquidacion_mes')
    )

    # 4. Cálculo de Estadísticas (Avance)
    total_trabajadores = trabajadores_qs.count()
    calculados_count = 0
    
    lista_trabajadores = []
    
    for t in trabajadores_qs:
        # Al usar prefetch con to_attr, el resultado es una lista. 
        # Si tiene elementos, es que ya se calculó.
        liq = t.liquidacion_mes[0] if t.liquidacion_mes else None
        
        if liq:
            calculados_count += 1
            
        lista_trabajadores.append({
            'obj': t,
            'liquidacion': liq, # Pasamos el objeto liquidación o None
            'tiene_calculo': bool(liq)
        })

    context = {
        'trabajadores_lista': lista_trabajadores,
        'periodo_str': periodo_str,
        'periodo_value': periodo_value,
        'total_count': total_trabajadores,
        'calculados_count': calculados_count,
    }
    
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/nomina_mensual.html', context)

def historial(request):
    """
    Muestra el historial de liquidaciones FILTRADO POR MES.
    """
    # 1. Gestión de Fechas (Igual que en nómina)
    fecha_get = request.GET.get('fecha') # Formato YYYY-MM
    hoy = timezone.now()
    
    if fecha_get:
        try:
            anio = int(fecha_get.split('-')[0])
            mes = int(fecha_get.split('-')[1])
            fecha_seleccionada = date(anio, mes, 1)
        except ValueError:
            fecha_seleccionada = hoy.date()
    else:
        fecha_seleccionada = hoy.date()
        
    periodo_value = fecha_seleccionada.strftime('%Y-%m')
    periodo_str = fecha_seleccionada.strftime('%B %Y').title()

    # 2. Filtrar Liquidaciones por el periodo seleccionado
    liquidaciones = Remuneracion.objects.filter(
        periodo=periodo_value
    ).select_related('trabajador').order_by('trabajador__nombre')
    
    # 3. Calcular Totales para mostrar resumen
    total_pagado = liquidaciones.aggregate(Sum('sueldo_liquido'))['sueldo_liquido__sum'] or 0
    total_registros = liquidaciones.count()

    context = {
        'liquidaciones': liquidaciones,
        'periodo_str': periodo_str,
        'periodo_value': periodo_value,
        'total_pagado': total_pagado,
        'total_registros': total_registros,
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_historial/remuneraciones_historial.html', context)

# ==========================================
# 2. UTILIDADES
# ==========================================

def format_currency_cl(value, decimal_places=0):
    """
    Formatea un número a string con punto de miles y coma decimal.
    """
    if value is None: return ""
    try:
        if decimal_places == 0:
            formatted = f"{int(value):,}".replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
        else:
            formatted = f"{value:,.{decimal_places}f}".replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
        return formatted
    except (ValueError, TypeError):
        return str(value)

def clean_currency(value):
    """Limpia cadenas de texto de símbolos de moneda, puntos de miles y comas para convertir a float."""
    if isinstance(value, str):
        value = value.replace('$', '').replace('.', '').replace(',', '.')
        try:
            return float(value)
        except ValueError:
            return 0.0
    return float(value) if value is not None else 0.0

def get_parametros_calculo():
    """Carga todos los parámetros necesarios en un solo diccionario."""
    config_raw = {c.clave: Decimal(str(c.valor)) for c in ConfiguracionGlobal.objects.all().iterator()}
    
    # Utilizamos el valor UF como tope de referencia para simplificar (81.6 UF)
    valor_uf = config_raw.get('valor_uf', Decimal('36500')) 
    tope_imponible_clp = valor_uf * Decimal('81.6')
    
    # Tope gratificación: 4.75 ingresos mínimos
    sueldo_minimo = config_raw.get('sueldo_minimo', Decimal('460000'))
    tope_gratificacion = (sueldo_minimo * Decimal('4.75')) / Decimal('12')

    return {
        'sueldo_minimo': sueldo_minimo,
        'valor_uf': valor_uf,
        'valor_utm': config_raw.get('valor_utm', Decimal('64000')),
        'gratificacion_pct': config_raw.get('gratificacion_legal_pct', Decimal('25')) / 100,
        'tope_gratificacion': tope_gratificacion,
        'seguro_cesantia_pct': config_raw.get('seguro_cesantia_pct', Decimal('0.6')) / 100,
        'tope_imponible': tope_imponible_clp.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
        'tramos_af': AsignacionFamiliarConfig.objects.all().order_by('ingreso_tope'),
    }

# ==========================================
# 3. CONFIGURACIÓN Y PARÁMETROS
# ==========================================

def parametros(request):
    """
    Vista principal que carga toda la configuración.
    """
    afps = AfpConfig.objects.all().order_by('nombre')
    salud = SaludConfig.objects.all().order_by('nombre')
    tramos = AsignacionFamiliarConfig.objects.all().order_by('ingreso_tope')
    
    config_global_qs = ConfiguracionGlobal.objects.all()
    config_raw = {c.clave: c.valor for c in config_global_qs}

    if not tramos.exists():
        AsignacionFamiliarConfig.objects.create(tramo='A', ingreso_tope=0, monto_por_carga=0)
        AsignacionFamiliarConfig.objects.create(tramo='B', ingreso_tope=0, monto_por_carga=0)
        AsignacionFamiliarConfig.objects.create(tramo='C', ingreso_tope=0, monto_por_carga=0)
        AsignacionFamiliarConfig.objects.create(tramo='D', ingreso_tope=999999999, monto_por_carga=0)
        tramos = AsignacionFamiliarConfig.objects.all().order_by('ingreso_tope')

    config_formato = {
        'gratificacion_legal_pct': config_raw.get('gratificacion_legal_pct', 0.0),
        'seguro_cesantia_pct': config_raw.get('seguro_cesantia_pct', 0.0),
    }

    context = {
        'afps': afps,
        'salud': salud,
        'tramos': tramos,
        'config': config_formato,
    }
    
    return render(request, 'gestion/remuneraciones/remuneraciones_parametros/remuneraciones_parametros.html', context)

def actualizar_indicador_singular(request):
    if request.method == 'POST':
        clave = request.POST.get('clave') 
        valor = request.POST.get('valor')
        descripcion = request.POST.get('descripcion', 'Indicador')

        if clave and valor:
            try:
                val_float = float(valor.replace(',', '.'))
                ConfiguracionGlobal.objects.update_or_create(
                    clave=clave,
                    defaults={'valor': val_float, 'descripcion': descripcion}
                )
            except (ValueError, TypeError):
                pass

    return redirect('parametros_remuneraciones')

def crear_entidad_previsional(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo_entidad') # 'afp' o 'salud'
        nombre = request.POST.get('nombre')
        tasa = request.POST.get('tasa')

        if nombre and tasa:
            try:
                tasa_float = float(tasa.replace(',', '.'))
                if tipo == 'afp':
                    AfpConfig.objects.create(nombre=nombre, tasa=tasa_float)
                elif tipo == 'salud':
                    SaludConfig.objects.create(nombre=nombre, tasa=tasa_float)
            except ValueError:
                pass 

    return redirect('parametros_remuneraciones')

def editar_entidad_previsional(request, tipo, id):
    if request.method == 'POST':
        if tipo == 'afp':
            instance = get_object_or_404(AfpConfig, id=id)
        elif tipo == 'salud':
            instance = get_object_or_404(SaludConfig, id=id)
        else:
            return redirect('parametros_remuneraciones')

        nombre_nuevo = request.POST.get('nombre')
        tasa_nueva = request.POST.get('tasa')

        if nombre_nuevo and tasa_nueva:
            try:
                instance.nombre = nombre_nuevo
                instance.tasa = float(tasa_nueva.replace(',', '.'))
                instance.save()
            except ValueError:
                pass
            
    return redirect('parametros_remuneraciones')

def eliminar_entidad_previsional(request, tipo, id):
    if request.method == 'POST':
        if tipo == 'afp':
            instance = get_object_or_404(AfpConfig, id=id)
            instance.delete()
        elif tipo == 'salud':
            instance = get_object_or_404(SaludConfig, id=id)
            instance.delete()
            
    return redirect('parametros_remuneraciones')

def editar_tramos_asignacion(request):
    if request.method == 'POST':
        tramos_qs = AsignacionFamiliarConfig.objects.all()
        try:
            for tramo in tramos_qs:
                tope_raw = request.POST.get(f'tope_{tramo.tramo}', '0')
                monto_raw = request.POST.get(f'monto_{tramo.tramo}', '0')
                
                tramo.ingreso_tope = int(clean_currency(tope_raw))
                tramo.monto_por_carga = int(clean_currency(monto_raw))
                tramo.save()
        except Exception:
            pass
            
    return redirect('parametros_remuneraciones')


# ==========================================
# 4. LÓGICA CORE: CÁLCULO DE SUELDO
# ==========================================

def calcular_sueldo(request, id):
    """
    Calcula el sueldo soportando múltiples listas dinámicas.
    """
    # 1. Determinar el Objeto Base y Periodo
    try:
        # Intenta cargar liquidación existente (Recalcular)
        liquidacion_existente = Remuneracion.objects.select_related('trabajador', 'trabajador__afp', 'trabajador__salud').get(id=id)
        trabajador = liquidacion_existente.trabajador
        periodo_actual = liquidacion_existente.periodo
        
    except Remuneracion.DoesNotExist:
        # Nuevo cálculo (ID es de Trabajador)
        trabajador = get_object_or_404(Trabajador.objects.select_related('afp', 'salud'), id=id)
        liquidacion_existente = None
        
        # --- Recuperar fecha desde URL o usar hoy ---
        periodo_get = request.GET.get('periodo')
        if periodo_get:
            periodo_actual = periodo_get
        else:
            periodo_actual = date.today().strftime('%Y-%m')

    p = get_parametros_calculo()
    
    # 3. Lógica POST: Calcular y Guardar
    if request.method == 'POST':
        
        dias_trabajados = int(request.POST.get('dias') or 30)
        horas_extras = float(request.POST.get('horas_extras') or 0)
        anticipo_fijo = clean_currency(request.POST.get('anticipo', 0))
        abono_faltante = clean_currency(request.POST.get('abono_faltante', 0))
        
        # --- PROCESAMIENTO DE LISTAS DINÁMICAS ---
        
        # Haberes
        json_haberes = request.POST.get('detalle_haberes', '[]')
        try:
            lista_haberes = json.loads(json_haberes)
            total_haberes_imponibles = sum(int(b.get('monto', 0) or 0) for b in lista_haberes)
        except:
            lista_haberes = []
            total_haberes_imponibles = 0

        # Asignaciones
        json_asignaciones = request.POST.get('detalle_asignaciones', '[]')
        try:
            lista_asignaciones = json.loads(json_asignaciones)
            total_asignaciones_manuales = sum(int(b.get('monto', 0) or 0) for b in lista_asignaciones)
        except:
            lista_asignaciones = []
            total_asignaciones_manuales = 0

        # Bonos No Imponibles
        json_bonos = request.POST.get('detalle_bonos', '[]')
        try:
            lista_bonos = json.loads(json_bonos)
            total_no_imponibles = sum(int(b.get('monto', 0) or 0) for b in lista_bonos)
        except:
            lista_bonos = []
            total_no_imponibles = 0

        # Descuentos Simples (Otros)
        json_descuentos = request.POST.get('detalle_descuentos', '[]')
        try:
            lista_descuentos = json.loads(json_descuentos)
            total_otros_descuentos = sum(int(b.get('monto', 0) or 0) for b in lista_descuentos)
        except:
            lista_descuentos = []
            total_otros_descuentos = 0
            
        # --- NUEVO: Descuentos Legales Variables (% o $) ---
        json_descuentos_legales = request.POST.get('detalle_descuentos_legales', '[]')
        try:
            # Parseamos, pero el cálculo se hace más abajo una vez tengamos el Imponible
            lista_descuentos_legales_raw = json.loads(json_descuentos_legales)
        except:
            lista_descuentos_legales_raw = []

        
        # --- CÁLCULOS MATEMÁTICOS ---
        
        # A. Base Imponible Parcial (Para calcular Gratificación)
        sueldo_base_dia = Decimal(trabajador.sueldo_base) / Decimal(30)
        sueldo_base_prop = (sueldo_base_dia * Decimal(dias_trabajados)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        valor_hora_extra = Decimal(trabajador.valor_hora_extra)
        monto_horas_extras = (Decimal(horas_extras) * valor_hora_extra).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        base_pre_gratificacion = sueldo_base_prop + monto_horas_extras + Decimal(total_haberes_imponibles)
        
        # B. Gratificación Legal (Con Tope 4.75 IMM)
        grat_teorica = base_pre_gratificacion * p['gratificacion_pct']
        tope_grat = p.get('tope_gratificacion', Decimal('99999999'))
        
        gratificacion_bruta = min(grat_teorica, tope_grat).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        # C. SUELDO IMPONIBLE FINAL (Base para Descuentos Legales)
        sueldo_imponible = base_pre_gratificacion + gratificacion_bruta
        
        # Tope Imponible
        base_calculo_descuentos = min(sueldo_imponible, p['tope_imponible'])
        
        # --- PROCESAR LISTA DESCUENTOS LEGALES VARIABLES ---
        lista_descuentos_legales_procesada = []
        total_descuentos_legales_extra = Decimal(0)
        
        for item in lista_descuentos_legales_raw:
            try:
                tipo = item.get('tipo', '$') # '$' o '%'
                valor_input = Decimal(str(item.get('valor', 0) or 0))
                desc = item.get('desc', '')
                
                monto_final = Decimal(0)
                if tipo == '%':
                    monto_final = (sueldo_imponible * (valor_input / Decimal(100))).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                else:
                    monto_final = valor_input
                
                total_descuentos_legales_extra += monto_final
                
                lista_descuentos_legales_procesada.append({
                    'desc': desc,
                    'tipo': tipo,
                    'valor': float(valor_input),
                    'monto_calculado': int(monto_final)
                })
            except:
                continue

        # D. Descuentos Legales Estándar
        monto_afp = (base_calculo_descuentos * Decimal(trabajador.afp.tasa) / Decimal(100)).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if trabajador.afp else Decimal('0')
        monto_salud = (base_calculo_descuentos * Decimal(trabajador.salud.tasa) / Decimal(100)).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if trabajador.salud else Decimal('0')
        
        # Seguro Cesantía
        tasa_cesantia_aplicada = p['seguro_cesantia_pct']
        cumplio_11_anos = False
        if trabajador.fecha_ingreso:
            hoy = date.today()
            antiguedad_anos = hoy.year - trabajador.fecha_ingreso.year - ((hoy.month, hoy.day) < (trabajador.fecha_ingreso.month, trabajador.fecha_ingreso.day))
            if antiguedad_anos >= 11:
                tasa_cesantia_aplicada = Decimal('0')
                cumplio_11_anos = True

        monto_seguro_cesantia = (base_calculo_descuentos * tasa_cesantia_aplicada).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        # E. Asignación Familiar
        monto_asignacion_automatica = Decimal('0')
        tramo_letra_aplicado = None
        tramo_monto_unitario = 0

        if trabajador.tiene_asignacion_familiar and trabajador.cargas_familiares > 0:
            for tramo in p['tramos_af']:
                if sueldo_imponible <= Decimal(tramo.ingreso_tope):
                    tramo_letra_aplicado = tramo.tramo
                    tramo_monto_unitario = tramo.monto_por_carga
                    monto_asignacion_automatica = Decimal(tramo.monto_por_carga) * Decimal(trabajador.cargas_familiares)
                    break
        
        monto_asignacion_familiar_total = monto_asignacion_automatica + Decimal(total_asignaciones_manuales)

        # --- IMPUESTO ÚNICO ---
        afecto_impuesto = sueldo_imponible - (monto_afp + monto_salud + monto_seguro_cesantia)
        
        if afecto_impuesto > 0:
            impuesto_determinado = afecto_impuesto * Decimal('0.04') # Simplificación tramo bajo
            rebaja_impuesto = Decimal('37218.42') # Valor aprox, debería ser paramétrico
            monto_impuesto_final = max(Decimal('0'), impuesto_determinado - rebaja_impuesto)
            monto_impuesto_final = monto_impuesto_final.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        else:
            monto_impuesto_final = Decimal('0')

        # --- FALTANTE DE CAJA ---
        monto_faltante_automatico = Decimal('0')
        if trabajador.filtro_trabajador:
            try:
                anio_liq, mes_liq = map(int, periodo_actual.split('-'))
                balance_mes = RendicionDiaria.objects.filter(
                    trabajador=trabajador,
                    fecha__year=anio_liq,
                    fecha__month=mes_liq
                ).aggregate(Sum('diferencia'))['diferencia__sum'] or 0
                
                if balance_mes < 0:
                    monto_faltante_automatico = Decimal(abs(balance_mes))
            except:
                pass
        
        monto_faltante_final = max(Decimal(0), monto_faltante_automatico - Decimal(abono_faltante))
        monto_faltante_final = monto_faltante_final.quantize(Decimal('1'), rounding=ROUND_HALF_UP)


        # F. Total de Haberes
        total_haberes = sueldo_imponible + monto_asignacion_familiar_total + Decimal(total_no_imponibles)
        
        # G. Total Descuentos
        total_descuentos_legales = monto_afp + monto_salud + monto_seguro_cesantia + monto_impuesto_final
        
        # Sumamos los legales extra que calculamos arriba
        total_descuentos = total_descuentos_legales + total_descuentos_legales_extra + Decimal(anticipo_fijo) + Decimal(total_otros_descuentos) + monto_faltante_final
        
        # H. Sueldo Líquido
        sueldo_liquido = total_haberes - total_descuentos
        
        detalle_calculo = {
            'sueldo_base_original': str(trabajador.sueldo_base),
            'sueldo_base_proporcional': str(sueldo_base_prop),
            'dias_trabajados': dias_trabajados,
            'horas_extras': horas_extras,
            'monto_horas_extras': str(monto_horas_extras),
            
            'lista_haberes': lista_haberes,
            'lista_asignaciones': lista_asignaciones,
            'lista_bonos': lista_bonos,
            'lista_descuentos': lista_descuentos,
            # Nueva lista guardada con los montos calculados
            'lista_descuentos_legales': lista_descuentos_legales_procesada, 
            
            'monto_anticipo': str(anticipo_fijo),
            
            'faltante_automatico_original': str(monto_faltante_automatico),
            'abono_faltante': str(abono_faltante),
            
            'monto_otros_haberes_total': str(Decimal(total_no_imponibles)),
            
            'asignacion_familiar': str(monto_asignacion_automatica),
            'monto_asignaciones_total': str(monto_asignacion_familiar_total),
            
            'monto_otros_descuentos_total': str(Decimal(total_otros_descuentos)),

            'sueldo_imponible': str(sueldo_imponible), 
            'gratificacion': str(gratificacion_bruta),
            'afp_nombre': trabajador.afp.nombre if trabajador.afp else 'N/A',
            'afp_tasa': float(trabajador.afp.tasa) if trabajador.afp else 0.0,
            'monto_afp': str(monto_afp),
            'salud_nombre': trabajador.salud.nombre if trabajador.salud else 'N/A',
            'salud_tasa': float(trabajador.salud.tasa) if trabajador.salud else 0.0,
            'monto_salud': str(monto_salud),
            'monto_seguro_cesantia': str(monto_seguro_cesantia),
            'cumplio_11_anos': cumplio_11_anos,

            'tramo_letra': tramo_letra_aplicado,
            'tramo_monto_unitario': str(tramo_monto_unitario),

            'monto_impuesto': str(monto_impuesto_final),
            'afecto_impuesto': str(afecto_impuesto),
            'impuesto_determinado': str(impuesto_determinado if afecto_impuesto > 0 else 0),
            
            'monto_faltante': str(monto_faltante_final),

            'total_haberes_calculado': str(total_haberes),
            'total_descuentos_calculado': str(total_descuentos),
            'periodo': periodo_actual,
            'rut': trabajador.rut
        }
        
        defaults = {
            'sueldo_liquido': int(sueldo_liquido),
            'total_haberes': int(total_haberes),
            'total_descuentos': int(total_descuentos),
            'monto_impuesto': int(monto_impuesto_final),
            'monto_faltante': int(monto_faltante_final),
            'detalle_json': detalle_calculo
        }
        
        if liquidacion_existente:
            for key, value in defaults.items():
                setattr(liquidacion_existente, key, value)
            liquidacion_existente.save()
            id_final = liquidacion_existente.id
        else:
            nueva = Remuneracion.objects.create(trabajador=trabajador, periodo=periodo_actual, **defaults)
            id_final = nueva.id
        
        return redirect('ver_liquidacion', id=id_final)
        
    # 4. LÓGICA GET: Cargar datos para el formulario
    
    initial_dias = 30
    initial_horas = 0
    initial_anticipo = 0
    initial_faltante_auto = 0
    initial_abono_faltante = 0

    l_haberes = []
    l_asignaciones = []
    l_bonos = []
    l_descuentos = []
    l_descuentos_legales = [] # Nueva lista

    if liquidacion_existente:
        detalle = liquidacion_existente.detalle_json
        initial_dias = detalle.get('dias_trabajados', 30)
        initial_horas = detalle.get('horas_extras', 0)
        try: initial_anticipo = int(float(detalle.get('monto_anticipo', 0)))
        except: initial_anticipo = 0
        
        try: initial_faltante_auto = int(float(detalle.get('faltante_automatico_original', 0)))
        except: initial_faltante_auto = 0
        
        try: initial_abono_faltante = int(float(detalle.get('abono_faltante', 0)))
        except: initial_abono_faltante = 0

        l_haberes = detalle.get('lista_haberes', [])
        l_asignaciones = detalle.get('lista_asignaciones', [])
        l_bonos = detalle.get('lista_bonos', [])
        l_descuentos = detalle.get('lista_descuentos', [])
        l_descuentos_legales = detalle.get('lista_descuentos_legales', [])
    
    else:
        if trabajador.filtro_trabajador:
            try:
                anio_liq, mes_liq = map(int, periodo_actual.split('-'))
                balance_mes = RendicionDiaria.objects.filter(
                    trabajador=trabajador,
                    fecha__year=anio_liq,
                    fecha__month=mes_liq
                ).aggregate(Sum('diferencia'))['diferencia__sum'] or 0
                
                if balance_mes < 0:
                    initial_faltante_auto = abs(balance_mes)
            except:
                initial_faltante_auto = 0

    initial_data = {
        'dias': initial_dias,
        'horas_extras': initial_horas,
        'anticipo': initial_anticipo,
        'abono_faltante': initial_abono_faltante,
        'json_haberes': json.dumps(l_haberes),
        'json_asignaciones': json.dumps(l_asignaciones),
        'json_bonos': json.dumps(l_bonos),
        'json_descuentos': json.dumps(l_descuentos),
        'json_descuentos_legales': json.dumps(l_descuentos_legales), # Pasamos la lista al template
    }
    
    # Parámetros JS para cálculo en vivo de tope gratificación
    tope_grat_js = p.get('tope_gratificacion', 0)

    context = {
        'trabajador': trabajador,
        'periodo_str': date(int(periodo_actual[:4]), int(periodo_actual[5:]), 1).strftime('%B %Y').upper(),
        'periodo_mes_nombre': date(int(periodo_actual[:4]), int(periodo_actual[5:]), 1).strftime('%B'),
        'initial': initial_data,
        'faltante_automatico': initial_faltante_auto,
        'tope_grat_anual': tope_grat_js
    }
    
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/formulario_calculo.html', context)

def ver_liquidacion(request, id):
    """
    Muestra el detalle de una liquidación calculada o guardada.
    """
    liquidacion = get_object_or_404(Remuneracion.objects.select_related('trabajador'), id=id)
    detalle = liquidacion.detalle_json
    
    detalle_formateado = {}
    for key, value in detalle.items():
        if key in ['afp_tasa', 'salud_tasa', 'horas_extras']:
            try:
                detalle_formateado[key] = float(value)
            except (ValueError, TypeError):
                detalle_formateado[key] = 0.0
        else:
            try:
                detalle_formateado[key] = int(float(value))
            except (ValueError, TypeError):
                detalle_formateado[key] = value

    # Rut
    rut_raw = liquidacion.trabajador.rut
    rut_formateado = rut_raw
    if rut_raw and len(rut_raw) > 1:
        rut_limpio = rut_raw.replace('.', '').replace('-', '').strip()
        if len(rut_limpio) > 1:
            cuerpo = rut_limpio[:-1]
            dv = rut_limpio[-1].upper()
            try:
                cuerpo_fmt = "{:,}".format(int(cuerpo)).replace(',', '.')
                rut_formateado = f"{cuerpo_fmt}-{dv}"
            except ValueError:
                rut_formateado = rut_raw

    # Bodega
    bodega_mostrar = liquidacion.trabajador.bodega_asignada
    if bodega_mostrar == 'Ambos':
        bodega_mostrar = liquidacion.trabajador.bodega_facturacion
    
    nombre_bodega = "Desconocida"
    if bodega_mostrar == '1221': nombre_bodega = "Bodega 1221 (Manuel Peñafiel)"
    elif bodega_mostrar == '1225': nombre_bodega = "Bodega 1225 (David Perry)"

    context = {
        'liquidacion': liquidacion,
        't': liquidacion.trabajador,
        'rut_fmt': rut_formateado,
        'bodega_real': nombre_bodega,
        'detalle': detalle_formateado,
        'total_haberes_brutos': liquidacion.total_haberes,
        'periodo_str': date(int(detalle_formateado['periodo'][:4]), int(detalle_formateado['periodo'][5:]), 1).strftime('%B %Y').upper(),
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/liquidacion_detalle.html', context)

# ==========================================
# 5. EXPORTACIÓN EXCEL
# ==========================================

def vista_excel_simulacion(request, id):
    """
    Muestra la vista previa del Excel.
    """
    liquidacion = get_object_or_404(Remuneracion, id=id)
    plantilla_obj = PlantillaLiquidacion.objects.first() 

    if request.method == 'POST' and 'plantilla_excel' in request.FILES:
        archivo = request.FILES['plantilla_excel']
        if plantilla_obj:
            plantilla_obj.archivo = archivo
            plantilla_obj.save()
        else:
            PlantillaLiquidacion.objects.create(archivo=archivo)
        return redirect('vista_excel_simulacion', id=id)

    preview_html = None
    error_msg = None

    if plantilla_obj:
        try:
            wb = openpyxl.load_workbook(plantilla_obj.archivo.path, data_only=True)
            ws = wb.active
            html = '<table class="table table-bordered table-sm text-center" style="font-size: 10px;">'
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12):
                html += '<tr>'
                for cell in row:
                    val = cell.value if cell.value is not None else ""
                    bg = "bg-light" if val else ""
                    html += f'<td class="{bg}">{val}</td>'
                html += '</tr>'
            html += '</table>'
            preview_html = html
        except Exception as e:
            error_msg = f"No se pudo generar la previsualización: {str(e)}"

    context = {
        'liquidacion': liquidacion,
        'plantilla_existe': (plantilla_obj is not None),
        'plantilla_fecha': plantilla_obj.updated_at if plantilla_obj else None,
        'preview_html': preview_html,
        'error_msg': error_msg
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_internos/remuneracion_excel.html', context)

def exportar_liquidacion_excel(request, id):
    liquidacion = get_object_or_404(Remuneracion, id=id)
    plantilla_obj = PlantillaLiquidacion.objects.first()

    if not plantilla_obj:
        return redirect('vista_excel_simulacion', id=id)

    trabajador = liquidacion.trabajador
    detalle = liquidacion.detalle_json

    try:
        wb = openpyxl.load_workbook(plantilla_obj.archivo.path)
        ws = wb.active
    except Exception:
        return HttpResponse("Error al abrir la plantilla guardada.", status=500)

    try:
        y, m = liquidacion.periodo.split('-')
        fecha_obj = date(int(y), int(m), 1)
        mes_nombre = fecha_obj.strftime('%B').title()
        anio_str = y
    except:
        mes_nombre = "---"
        anio_str = "----"

    def fmt_peso(valor):
        try:
            val = int(float(valor))
            return f"$ {val:,.0f}".replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
        except:
            return "$ 0"

    sueldo_liquido = fmt_peso(liquidacion.sueldo_liquido)
    dias_trabajados = detalle.get('dias_trabajados', 30)
    sueldo_diario = fmt_peso(int(trabajador.sueldo_base) / 30)
    horas_extras_str = str(detalle.get('horas_extras', 0)).replace('.', ',')
    fecha_ingreso = trabajador.fecha_ingreso.strftime('%d-%m-%Y') if trabajador.fecha_ingreso else ""
    
    rut_fmt = trabajador.rut or ""
    bodega_nombre = trabajador.bodega_asignada
    if bodega_nombre == '1221': bodega_nombre = 'Manuel Peñafiel (1221)'
    elif bodega_nombre == '1225': bodega_nombre = 'David Perry (1225)'

    font_bold = Font(name='Arial', size=10, bold=True)
    font_normal = Font(name='Arial', size=10)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    def celda(coord, valor, font, align):
        c = ws[coord]
        c.value = valor
        c.font = font
        c.alignment = align

    celda('I2', mes_nombre, font_bold, align_center)
    celda('J2', anio_str, font_bold, align_center)
    celda('I4', sueldo_liquido, font_bold, align_center)
    celda('E8', trabajador.nombre, font_normal, align_left)
    celda('E9', trabajador.cargo or "Operario", font_bold, align_left)
    celda('E10', bodega_nombre, font_normal, align_left)
    celda('E11', dias_trabajados, font_normal, align_center)
    celda('E12', "25%", font_normal, align_center)
    celda('I8', rut_fmt, font_normal, align_left)
    celda('I10', fecha_ingreso, font_normal, align_right)
    celda('I11', sueldo_diario, font_normal, align_center)
    celda('I12', horas_extras_str, font_normal, align_right)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Liquidacion_{trabajador.nombre.replace(' ', '_')}_{mes_nombre}.xlsx"
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ==========================================
# 6. REMUNERACIONES EXTERNOS
# ==========================================

def nomina_externos(request):
    """
    Lista de trabajadores con contrato 'EXTERNO'.
    """
    fecha_get = request.GET.get('fecha')
    if fecha_get:
        fecha_dt = datetime.strptime(fecha_get, '%Y-%m')
    else:
        fecha_dt = timezone.now()

    mes = fecha_dt.month
    anio = fecha_dt.year

    trabajadores = Trabajador.objects.filter(tipo='EXTERNO', activo=True).order_by('nombre')
    
    # Diccionario para saber quién ya tiene factura procesada
    procesados = RemuneracionExterna.objects.filter(mes=mes, anio=anio).values_list('trabajador_id', flat=True)

    context = {
        'trabajadores': trabajadores,
        'mes': mes,
        'anio': anio,
        'procesados': procesados,
        'fecha_actual': fecha_dt.strftime('%Y-%m'),
        'mes_nombre': calendar.month_name[mes].capitalize()
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_externos/sueldos_externos.html', context)


def calcular_remuneracion_externa(request, trabajador_id):
    """
    Calcula la remuneración de un trabajador externo.
    """
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    mes = int(request.GET.get('mes', timezone.now().month))
    anio = int(request.GET.get('anio', timezone.now().year))

    tarifas = TarifaComision.objects.last()
    if not tarifas:
        tarifas = TarifaComision(
            tarifa_5kg=0, tarifa_11kg=0, tarifa_15kg=0, tarifa_45kg=0,
            tarifa_cat_5kg=0, tarifa_cat_15kg=0, tarifa_ultra_15kg=0
        )

    # 1. Recuperar Ventas y Faltantes de RendicionDiaria
    rendiciones = RendicionDiaria.objects.filter(
        trabajador=trabajador,
        fecha__month=mes,
        fecha__year=anio
    )

    conteo_cilindros = {
        '5kg': 0, '11kg': 0, '15kg': 0, '45kg': 0, 
        'cat5kg': 0, 'cat15kg': 0, 'ultra15': 0
    }

    faltante_acumulado = 0

    for r in rendiciones:
        conteo_cilindros['5kg'] += (r.gas_5kg or 0)
        conteo_cilindros['11kg'] += (r.gas_11kg or 0)
        conteo_cilindros['15kg'] += (r.gas_15kg or 0)
        conteo_cilindros['45kg'] += (r.gas_45kg or 0)
        
        conteo_cilindros['cat5kg'] += (r.gasc_5kg or 0)
        conteo_cilindros['cat15kg'] += (r.gasc_15kg or 0)
        conteo_cilindros['ultra15'] += (r.gas_ultra_15kg or 0)
        
        if r.diferencia < 0:
            faltante_acumulado += abs(r.diferencia)

    total_pago_cilindros = (
        conteo_cilindros['5kg'] * tarifas.tarifa_5kg +
        conteo_cilindros['11kg'] * tarifas.tarifa_11kg +
        conteo_cilindros['15kg'] * tarifas.tarifa_15kg +
        conteo_cilindros['45kg'] * tarifas.tarifa_45kg +
        
        conteo_cilindros['cat5kg'] * tarifas.tarifa_cat_5kg +
        conteo_cilindros['cat15kg'] * tarifas.tarifa_cat_15kg +
        conteo_cilindros['ultra15'] * tarifas.tarifa_ultra_15kg
    )

    if request.method == 'POST':
        nro_factura = request.POST.get('nro_factura')
        asistencia = int(request.POST.get('asistencia_tecnica') or 0)
        prestamo = int(request.POST.get('prestamo') or 0)
        anticipo = int(request.POST.get('anticipo') or 0)
        faltante_final = int(request.POST.get('faltante') or 0)

        neto = total_pago_cilindros + asistencia
        iva = int(neto * 0.19)
        bruto = neto + iva
        
        total_pagar = bruto - (prestamo + anticipo + faltante_final)

        remu, created = RemuneracionExterna.objects.update_or_create(
            trabajador=trabajador, mes=mes, anio=anio,
            defaults={
                'nro_factura': nro_factura,
                'pago_cilindros': total_pago_cilindros,
                'asistencia_tecnica': asistencia,
                'subtotal_neto': neto,
                'iva': iva,
                'total_bruto': bruto,
                'prestamo': prestamo,
                'anticipo': anticipo,
                'faltante': faltante_final,
                'monto_total_pagar': total_pagar,
                'json_detalle_cilindros': json.dumps(conteo_cilindros)
            }
        )
        return redirect('detalle_remuneracion_externa', remu_id=remu.id)

    context = {
        'trabajador': trabajador,
        'mes': mes,
        'anio': anio,
        'pago_cilindros': total_pago_cilindros,
        'faltante_auto': int(faltante_acumulado),
        'conteo': conteo_cilindros
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_externos/calculo_externos.html', context)


def detalle_remuneracion_externa(request, remu_id):
    """
    Muestra el detalle final de la remuneración externa.
    """
    remu = get_object_or_404(RemuneracionExterna, id=remu_id)
    detalle_ventas = json.loads(remu.json_detalle_cilindros) if remu.json_detalle_cilindros else {}
    
    context = {
        'remu': remu,
        'detalle_ventas': detalle_ventas
    }
    return render(request, 'gestion/remuneraciones/remuneraciones_externos/detalle_externos.html', context)